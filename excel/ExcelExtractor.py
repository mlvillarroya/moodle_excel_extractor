"""Excel extractor class"""
import json
import openpyxl
from os import path as OSPath
from misc import ProjectPaths

constants_path = ProjectPaths.get_constants_path()
with open(constants_path, "r", encoding="utf8") as file:
    constants = json.load(file)

class ExcelExtractor:
    """Class to extract excel information into an array"""
    def __init__(self, filepath):
        try:
            self.__wb = openpyxl.load_workbook(filepath)
        except FileNotFoundError:
            raise FileNotFoundError("File not found. Please check the path")
        settings_constants = constants["settings_sheet"]
        try:
            settings_sheet = self.__wb[settings_constants["name"]]
        except KeyError:
            raise KeyError("A sheet named Settings is missing. Please, create one")
        subject_name_cell = settings_constants["data"]["subject_name"]
        chapter_name_cell = settings_constants["data"]["chapter_name"]
        subject_name = settings_sheet[subject_name_cell].value # type: ignore
        chapter_name = settings_sheet[chapter_name_cell].value # type: ignore
        self.__main_category = '/'.join(filter(None, (subject_name, chapter_name)))

    @property
    def main_category(self):
        """Property: main category"""
        return self.__main_category

    @property
    def workbook(self):
        """Property: workbook"""
        return self.__wb

    def extract_questions_from_workbook(self):
        """Method: extract questions from every sheet in workbook"""
        sheet_names = [constants[sheet]["name"] for sheet in constants.keys()]
        return {name:self.extract_questions_from_sheet(name) for name in sheet_names if name != "Settings"}

    def extract_questions_from_sheet(self,sheet_name):
        """Method: extract question from one sheet"""
        try:
            worksheet = self.__wb[sheet_name]
        except KeyError:
            return None
        if not self.__check_sheet_has_mandatory_columns(sheet_name):
            raise ValueError(f"Worksheet {sheet_name} has not the mandatory columns")
        headers = [cell.value for cell in self.__wb[sheet_name][1]]
        mandatory_columns_index = self.__get_mandatory_columns(worksheet)
        questions_array = []
        for row in worksheet.iter_rows(min_col=1,
                                       min_row=2,
                                       max_col=self.__get_sheet_width(worksheet), 
                                       max_row=self.__get_sheet_height(worksheet)):
            if self.__every_mandatory_field_is_filled(mandatory_columns_index, row):
                question_data = {headers[i]:cell.value for i, cell in enumerate(row)}
                questions_array.append(question_data)
        return questions_array

    def __get_sheet_width(self, sheet):
        return sheet.max_column

    def __get_sheet_height(self, sheet):
        return sheet.max_row

    def __get_mandatory_columns(self, sheet):
        first_row = sheet[1]
        mandatory_columns = [i for i, cell in enumerate(first_row) if cell.value is not None and "*" in cell.value]
        return mandatory_columns

    def __check_sheet_has_mandatory_columns(self, sheet_name):
        sheets_names = list(constants.keys())
        sheet_constants = [constants[k] for k in sheets_names if constants[k]["name"] == sheet_name][0]
        sheet_mandatory_headers = [value[1] for value in sheet_constants["cell_titles"].values() if "*" in value[1]]
        worksheet_headers = [cell.value for cell in self.__wb[sheet_name][1]]
        for header in sheet_mandatory_headers:
            if header not in worksheet_headers:
                return False
        return True

    def __every_mandatory_field_is_filled(self, mandatory_columns, row):
        for column in mandatory_columns:
            if row[column].value is None:
                return False
        return True
