from openpyxl.comments import Comment
from excel.ExcelCellStyling import ExcelCellStyling
import misc.Constants as CS

class SettingsSheet: 
    def configureSettingsSheet(ws, demoData = False):
        ws.title = CS.SETTINGS_SHEET_NAME
        ws['A1'] = CS.SETTINGS_SUBJECT_NAME_TITLE
        ws['B1'] = CS.SETTINGS_CHAPTER_NAME_TITLE
        ws['E1'] = CS.SETTINGS_MANDATORY_FIELD_LABEL
        for row in ws['A1':'B1']:
            for cell in row:
                ExcelCellStyling.ChangeCellAlignment(cell,'center','center')
                ExcelCellStyling.ChangeCellBackgroundAndTextColors(cell,'E7AA73','653159')
                ExcelCellStyling.adjustSheetColumnSize(ws,cell.column_letter,18)
                ExcelCellStyling.ApplyFullBorderToCell(cell)
        ExcelCellStyling.adjustSheetRowSize(ws,1,38)
        if demoData:
            ws['A2'] = "SMX_M03"
            ws['B2'] = "Spreadsheet"
        for row in ws['A2':'B2']:
            for cell in row:
                ExcelCellStyling.ChangeCellAlignment(cell,wrap_text=True)
                ExcelCellStyling.ApplyFullBorderToCell(cell)
