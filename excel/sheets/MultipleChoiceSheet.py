from openpyxl.comments import Comment
from excel.ExcelCellStyling import ExcelCellStyling
import misc.Constants as CS

class MultipleChoiceSheet: 
    def createMultipleChoiceSheet(wb, demoData = False):
        ws = wb.create_sheet(CS.MULTIPLE_CHOICE_SHEET_NAME)
        ws['A1'] = CS.MULTIPLE_CHOICE_QUESTION_TITLE
        ws['B1'] = CS.MULTIPLE_CHOICE_CORRECT_ANSWER_TITLE
        ws['C1'] = CS.MULTIPLE_CHOICE_INCORRECT_ANSWER_1_TITLE
        ws['D1'] = CS.MULTIPLE_CHOICE_INCORRECT_ANSWER_2_TITLE
        ws['E1'] = CS.MULTIPLE_CHOICE_INCORRECT_ANSWER_3_TITLE
        ws['F1'] = CS.MULTIPLE_CHOICE_INCORRECT_ANSWER_4_TITLE
        ws['G1'] = CS.MULTIPLE_CHOICE_BAD_ANSWER_POINTS_TITLE
        ws['G1'].comment = Comment(CS.MULTIPLE_CHOICE_BAD_ANSWER_POINTS_COMMENT,"admin",250,200)
        ws['H1'] = CS.MULTIPLE_CHOICE_FEEDBACK_TITLE
        ws['I1'] = CS.MULTIPLE_CHOICE_SUBCATEGORY_TITLE
        for row in ws['A1':'I1']:
            for cell in row:
                ExcelCellStyling.ChangeCellAlignment(cell,'center','center')
                ExcelCellStyling.ChangeCellBackgroundAndTextColors(cell,'E7AA73','653159')
                ExcelCellStyling.adjustSheetColumnSize(ws,cell.column_letter,18)
                ExcelCellStyling.ApplyFullBorderToCell(cell)
        ExcelCellStyling.adjustSheetRowSize(ws,1,38)
        if demoData:
            ws['A2'] = "Meaning of CPU?"
            ws['B2'] = "Control process unit"
            ws['C2'] = "Carpenter public uniform"
            ws['D2'] = "Car pen use"
            ws['E2'] = "Coconut private use"
            ws['F2'] = "Cardigan pickle uniform"
            ws['G2'] = "-30"
            ws['H2'] = "A central processing unit (CPU) is the electronic circuitry that executes instructions comprising a computer program."
            ws['I2'] = "Processing"
        for row in ws['A2':'I20']:
            for cell in row:
                ExcelCellStyling.ChangeCellAlignment(cell,wrap_text=True)
                ExcelCellStyling.ApplyFullBorderToCell(cell)
