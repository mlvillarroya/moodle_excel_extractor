from openpyxl.comments import Comment
from excel.ExcelCellStyling import ExcelCellStyling
import misc.Constants as CS

class OneAnswerSheet: 
    def createOneAnswerSheet(wb, demoData = False):
        ws = wb.create_sheet(CS.ONE_ANSWER_SHEET_NAME)
        ws['A1'] = CS.ONE_ANSWER_QUESTION_TITLE
        ws['B1'] = CS.ONE_ANSWER_CORRECT_ANSWER_TITLE
        ws['C1'] = CS.ONE_ANSWER_ALTERNATE_CORRECT_ANSWER_1_TITLE
        ws['D1'] = CS.ONE_ANSWER_ALTERNATE_CORRECT_ANSWER_2_TITLE
        ws['E1'] = CS.ONE_ANSWER_ALTERNATE_CORRECT_ANSWER_3_TITLE
        ws['F1'] = CS.ONE_ANSWER_ALTERNATE_CORRECT_ANSWER_4_TITLE
        ws['G1'] = CS.ONE_ANSWER_FEEDBACK_TITLE
        ws['H1'] = CS.ONE_ANSWER_SUBCATEGORY_TITLE
        for row in ws['A1':'H1']:
            for cell in row:
                ExcelCellStyling.ChangeCellAlignment(cell,'center','center')
                ExcelCellStyling.ChangeCellBackgroundAndTextColors(cell,'E7AA73','653159')
                ExcelCellStyling.adjustSheetColumnSize(ws,cell.column_letter,18)
                ExcelCellStyling.ApplyFullBorderToCell(cell)
        ExcelCellStyling.adjustSheetRowSize(ws,1,38)
        if demoData:
            ws['A2'] = "Color of one original Power Rangers"
            ws['B2'] = "Red"
            ws['C2'] = "Blue"
            ws['D2'] = "Yellow"
            ws['E2'] = "Pink"
            ws['F2'] = "Black"
            ws['G2'] = "The color of the original Rangers are red, blue, yellow, pink and black"
            ws['H2'] = "Television"
        for row in ws['A2':'H20']:
            for cell in row:
                ExcelCellStyling.ChangeCellAlignment(cell,wrap_text=True)
                ExcelCellStyling.ApplyFullBorderToCell(cell)
