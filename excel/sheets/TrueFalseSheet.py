from openpyxl.comments import Comment
from excel.ExcelCellStyling import ExcelCellStyling
import misc.Constants as CS

class TrueFalseSheet: 
    def createTrueFalseSheet(wb, demoData = False):
        ws = wb.create_sheet(CS.TRUE_FALSE_SHEET_NAME)
        ws['A1'] = CS.TRUE_FALSE_QUESTION_TITLE
        ws['B1'] = CS.TRUE_FALSE_ANSWER_TITLE
        ws['C1'] = CS.TRUE_FALSE_FEEDBACK_TITLE
        ws['D1'] = CS.TRUE_FALSE_SUBCATEGORY_TITLE
        for row in ws['A1':'D1']:
            for cell in row:
                ExcelCellStyling.ChangeCellAlignment(cell,'center','center')
                ExcelCellStyling.ChangeCellBackgroundAndTextColors(cell,'E7AA73','653159')
                ExcelCellStyling.adjustSheetColumnSize(ws,cell.column_letter,20)
                ExcelCellStyling.ApplyFullBorderToCell(cell)
        ExcelCellStyling.adjustSheetRowSize(ws,1,38)
        if demoData:
            ws['A2'] = "RAM memory is volatile"
            ws['B2'] = "T"
            ws['C2'] = "RAM is volatile memory, which means that the information temporarily stored in the module is erased when you restart or shut down your computer."
            ws['D2'] = "Processing"
        for row in ws['A2':'D20']:
            for cell in row:
                ExcelCellStyling.ChangeCellAlignment(cell,wrap_text=True)
                ExcelCellStyling.ApplyFullBorderToCell(cell)
