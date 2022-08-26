from openpyxl.comments import Comment
from excel.ExcelCellStyling import ExcelCellStyling
import misc.Constants as CS

class NumericSheet: 
    def createNumericSheet(wb, demoData = False):
        ws = wb.create_sheet(CS.NUMERIC_SHEET_NAME)
        ws['A1'] = CS.NUMERIC_QUESTION_TITLE
        ws['B1'] = CS.NUMERIC_CORRECT_VALUE_TITLE
        ws['C1'] = CS.NUMERIC_TOLERANCE_TITLE
        ws['D1'] = CS.NUMERIC_FEEDBACK_TITLE
        ws['E1'] = CS.NUMERIC_SUBCATEGORY_TITLE
        for row in ws['A1':'E1']:
            for cell in row:
                ExcelCellStyling.ChangeCellAlignment(cell,'center','center')
                ExcelCellStyling.ChangeCellBackgroundAndTextColors(cell,'E7AA73','653159')
                ExcelCellStyling.adjustSheetColumnSize(ws,cell.column_letter,18)
                ExcelCellStyling.ApplyFullBorderToCell(cell)
        ExcelCellStyling.adjustSheetRowSize(ws,1,38)
        if demoData:
            ws['A2'] = "Value of Pi?"
            ws['B2'] = "3.14"
            ws['C2'] = "0.01"
            ws['D2'] = "Value of Pi is 3.141592653589793238"
            ws['E2'] = "Math"
        for row in ws['A2':'E20']:
            for cell in row:
                ExcelCellStyling.ChangeCellAlignment(cell,wrap_text=True)
                ExcelCellStyling.ApplyFullBorderToCell(cell)
