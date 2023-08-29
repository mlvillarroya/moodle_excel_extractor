from pathlib import Path
import json
from os import remove as OSRemove
from os import path as OSPath
from enum import Enum
from openpyxl import Workbook
from openpyxl.comments import Comment
import misc.Constants as CS
from . import ExcelCellStyling

with open("static/excel_creation_constants.json", "r", encoding="utf8") as file:
    constants = json.load(file)
styling = ExcelCellStyling()

class ExcelMode(Enum):
    """Enum: Excel creation or extraction"""
    CREATE = 1
    EXTRACT = 2

class ExcelCreator:
    """Class to create excel files with example"""
    def __init__(self,
                 path = '',
                 excel_mode: ExcelMode = ExcelMode.CREATE,
                 create_multiple_choice = False,
                 create_true_false = False,
                 create_one_answer = False,
                 create_numeric = False,
                 demo_data = False):
        if excel_mode == ExcelMode.CREATE:
            self.__path = path if path != '' else Path().resolve().__str__()
            self.__filename = CS.FILENAME
            self.__wb = Workbook()
            self.__configure_settings_sheet(demo_data)
            if create_multiple_choice:
                self.__create_multiple_choice_sheet(demo_data)
            if create_one_answer:
                self.__create_one_answer_sheet(demo_data)
            if create_true_false:
                self.__create_true_false_sheet(demo_data)
            if create_numeric:
                self.__create_numeric_sheet(demo_data)

    @property
    def path(self):
        """Property: path"""
        return self.__path

    @property
    def filename(self):
        """Property: path"""
        return self.__filename

    @property
    def workbook(self):
        """Property: path"""
        return self.__wb

    def save_excel_file(self):
        """Function to save the workbook into a file"""
        self.__wb.save(OSPath.join(self.__path,self.__filename))

    def remove_excel_file(self):
        """Function to save the workbook into a file"""
        OSRemove(OSPath.join(self.__path,self.__filename))

    def __configure_settings_sheet(self, demo_data= False):
        settings_constants = constants["settings_sheet"]
        first_sheet = self.__wb.active
        self.__insert_cell_text_into_sheet(settings_constants, first_sheet)
        styling.first_row_adequation(first_sheet)
        if demo_data:
            self.__insert_demo_data_into_sheet(settings_constants, first_sheet)

    def __create_multiple_choice_sheet(self, demo_data= False):
        """Function to create multiple choice sheets"""
        multiple_choice_constants = constants["multiple_choice_sheet"]
        multiple_choice_sheet = self.__wb.create_sheet()
        self.__insert_cell_text_into_sheet(multiple_choice_constants, multiple_choice_sheet)
        styling.first_row_adequation(multiple_choice_sheet)
        if demo_data:
            self.__insert_demo_data_into_sheet(multiple_choice_constants, multiple_choice_sheet)

    def __create_one_answer_sheet(self, demo_data = False):
        one_answer_constants = constants["one_answer_sheet"]
        one_answer_sheet = self.__wb.create_sheet()
        self.__insert_cell_text_into_sheet(one_answer_constants, one_answer_sheet)
        styling.first_row_adequation(one_answer_sheet)
        if demo_data:
            self.__insert_demo_data_into_sheet(one_answer_constants, one_answer_sheet)

    def __create_true_false_sheet(self, demo_data = False):
        true_false_constants = constants["true_false_sheet"]
        true_false_sheet = self.__wb.create_sheet()
        self.__insert_cell_text_into_sheet(true_false_constants, true_false_sheet)
        styling.first_row_adequation(true_false_sheet)
        if demo_data:
            self.__insert_demo_data_into_sheet(true_false_constants, true_false_sheet)

    def __create_numeric_sheet(self, demo_data = False):
        numeric_constants = constants["numeric_sheet"]
        numeric_sheet = self.__wb.create_sheet()
        self.__insert_cell_text_into_sheet(numeric_constants, numeric_sheet)
        styling.first_row_adequation(numeric_sheet)
        if demo_data:
            self.__insert_demo_data_into_sheet(numeric_constants, numeric_sheet)

    def __insert_cell_text_into_sheet(self, constants_dict, sheet):
        sheet_name = constants_dict["name"]
        sheet.title = sheet_name
        sheet_cells_data = constants_dict["cells"]
        for coordinates, name in sheet_cells_data.values():
            sheet[coordinates] = name

    def __insert_demo_data_into_sheet(self, constants_dict, sheet):
        demo_data = constants_dict["demo_data"]
        for coordinates, name in demo_data.values():
            sheet[coordinates] = name
